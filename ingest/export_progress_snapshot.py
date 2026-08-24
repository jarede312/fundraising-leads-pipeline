"""
Ad hoc deliverable, not a pipeline phase: a father-in-law-readable snapshot of what the
database holds so far - one Excel workbook with a summary tab, the contacts found across
all sources, and the PTO/booster orgs matched to schools. Meant to be handed to a
non-technical stakeholder alongside the progress report, not consumed by any later phase.

Run: python -m ingest.export_progress_snapshot
Writes: data/out/progress_snapshot_<date>.xlsx
"""
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from . import config, db

HEADER_FILL = PatternFill("solid", fgColor="F1ECE2")
HEADER_FONT = Font(bold=True)


def _autosize(ws):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 2, 10), 48)


def _write_table(ws, headers, rows):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    _autosize(ws)


def main():
    conn = db.connect()
    cur = conn.cursor()

    wb = openpyxl.Workbook()

    # --- Summary ---
    cur.execute("SELECT count(*) FROM districts")
    n_districts = cur.fetchone()[0]
    cur.execute("SELECT count(*) FROM schools WHERE status = 'open'")
    n_schools = cur.fetchone()[0]
    cur.execute("SELECT count(*) FROM schools WHERE status = 'open' AND district_id IS NOT NULL")
    n_public = cur.fetchone()[0]
    cur.execute("SELECT count(*) FROM schools WHERE status = 'open' AND district_id IS NULL")
    n_private = cur.fetchone()[0]
    cur.execute("SELECT count(*) FROM contacts")
    n_contacts = cur.fetchone()[0]
    cur.execute("SELECT count(*) FILTER (WHERE email IS NOT NULL) FROM contacts")
    n_contacts_email = cur.fetchone()[0]
    cur.execute("SELECT count(DISTINCT school_id) FROM contacts")
    n_schools_with_contact = cur.fetchone()[0]
    cur.execute("SELECT count(DISTINCT ein) FROM nonprofit_orgs")
    n_orgs = cur.fetchone()[0]
    cur.execute("SELECT count(*) FROM school_org_links WHERE confirmed = true")
    n_orgs_confirmed = cur.fetchone()[0]
    cur.execute("SELECT count(*) FROM school_org_links WHERE confirmed = false")
    n_orgs_pending = cur.fetchone()[0]

    ws = wb.active
    ws.title = "Summary"
    _write_table(ws, ["Metric", "Count"], [
        ("Report date", date.today().isoformat()),
        ("School districts (North Dakota)", n_districts),
        ("Open schools", n_schools),
        ("  - public", n_public),
        ("  - private", n_private),
        ("Staff/PTO contacts on file", n_contacts),
        ("  - with an email address", n_contacts_email),
        ("Schools with at least one contact", n_schools_with_contact),
        ("PTO / booster club / school foundation orgs identified", n_orgs),
        ("  - confidently matched to a specific school", n_orgs_confirmed),
        ("  - found, but needs a human to confirm the match", n_orgs_pending),
    ])

    # --- Contacts ---
    cur.execute(
        """
        SELECT s.name, s.city, d.name, s.segment, c.role, c.role_detail,
               c.first_name, c.last_name, c.email, c.email_confidence, c.phone,
               ir.source, c.first_seen::date
        FROM contacts c
        JOIN schools s ON s.id = c.school_id
        LEFT JOIN districts d ON d.id = s.district_id
        LEFT JOIN ingest_runs ir ON ir.id = c.ingest_run_id
        ORDER BY s.name, c.role
        """
    )
    rows = cur.fetchall()
    ws2 = wb.create_sheet("Contacts")
    _write_table(ws2, [
        "School", "City", "District", "School type", "Role", "Title (as printed)",
        "First name", "Last name", "Email", "Email confidence", "Phone",
        "Source", "First seen",
    ], rows)

    # --- PTO / booster orgs ---
    cur.execute(
        """
        SELECT n.name, n.city, s.name, s.city, l.match_score, l.confirmed,
               n.filing_type, n.total_revenue, n.principal_officer_name, n.ein
        FROM school_org_links l
        JOIN nonprofit_orgs n ON n.ein = l.ein
        JOIN schools s ON s.id = l.school_id
        ORDER BY l.confirmed DESC, l.match_score DESC
        """
    )
    rows = cur.fetchall()
    ws3 = wb.create_sheet("PTO & Booster Orgs")
    _write_table(ws3, [
        "Organization", "Org city", "Matched school", "School city", "Match confidence",
        "Confirmed", "IRS filing type", "Total revenue ($)", "Principal officer", "EIN",
    ], rows)

    conn.close()

    out_path = config.OUT_DIR / f"progress_snapshot_{date.today().isoformat()}.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    main()
